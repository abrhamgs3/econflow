"""
econflow.ingestion.connectors.pwt -- Penn World Tables connector.

Downloads the Penn World Tables (PWT) panel dataset from the official
Groningen Growth and Development Centre (GGDC) release page.

Reference
---------
* PWT homepage: https://www.rug.nl/ggdc/productivity/pwt/
* PWT 10.01 data: https://dataverse.harvard.edu/dataset.xhtml?persistentId=doi:10.7910/DVN/61LOLE

Usage
-----
::

    from econflow.ingestion.connectors import PennWorldTablesConnector
    from econflow.ingestion.cache import CacheManager

    cache = CacheManager(".cache/econflow")
    conn = PennWorldTablesConnector(
        params={
            "version": "10.01",
            "variables": ["rgdpna", "emp", "avh", "rkna", "rtfpna"],
        },
        cache_manager=cache,
    )
    path, meta = conn.fetch()
"""

from __future__ import annotations

import csv
from pathlib import Path
from typing import Any

from econflow.ingestion.base import AbstractConnector, ConnectorError
from econflow.ingestion.metadata import DatasetMetadata
from econflow.ingestion.registry import register
from econflow.ingestion.validation import DataValidationConfig, DataValidationReport, DataValidator

# PWT release download URLs keyed by version string.
# The direct Dataverse file download links.
_PWT_URLS: dict[str, str] = {
    "10.01": (
        "https://dataverse.harvard.edu/api/access/datafile/"
        ":persistentId?persistentId=doi:10.7910/DVN/61LOLE/RWKJTZ"
    ),
    "10.0": (
        "https://dataverse.harvard.edu/api/access/datafile/"
        ":persistentId?persistentId=doi:10.7910/DVN/61LOLE/RWKJTZ"
    ),
}
_DEFAULT_VERSION = "10.01"
_CITATION = (
    "Feenstra, R.C., Inklaar, R., and Timmer, M.P. (2015). "
    "'The Next Generation of the Penn World Table.' "
    "American Economic Review, 105(10), pp.3150-82. "
    "Available for download at www.ggdc.net/pwt"
)


@register(
    "pwt",
    label="Penn World Tables",
    status="implemented",
    notes="PWT 10.01; Excel download from Harvard Dataverse; requires openpyxl",
)
class PennWorldTablesConnector(AbstractConnector):
    """
    Connector for the Penn World Tables (PWT).

    Downloads the official PWT Excel workbook from Harvard Dataverse,
    extracts the ``data`` sheet, optionally subsets to requested variables,
    and writes a tidy wide-format CSV with one row per (country, year).

    Parameters (``params`` dict keys)
    -----------------------------------
    version : str
        PWT version to download (e.g. ``"10.01"``).
        Supported versions: ``"10.0"``, ``"10.01"``.
        Defaults to ``"10.01"``.
    variables : list[str] | None
        List of PWT variable codes to include in the output CSV.
        ``None`` or omitted returns all variables.  Useful variable codes:

        * ``"rgdpna"``  — Real GDP at constant 2017 national prices (in mil. 2017 USD)
        * ``"emp"``     — Number of persons engaged (in millions)
        * ``"avh"``     — Average annual hours worked by persons engaged
        * ``"rkna"``    — Capital stock at constant 2017 national prices (in mil. 2017 USD)
        * ``"rtfpna"``  — TFP at constant national prices (2017=1)
        * ``"labsh"``   — Share of labour compensation in GDP at current national prices
    entity_col : str
        Column name for the country dimension in output.  Default ``"country"``.
    time_col : str
        Column name for the time dimension in output.  Default ``"year"``.
    timeout : int
        HTTP timeout in seconds.  Default 120 (large file).
    """

    _CITATION = _CITATION
    _VERSION = _DEFAULT_VERSION

    def __init__(
        self,
        params: dict[str, Any] | None = None,
        cache_manager: Any | None = None,
    ) -> None:
        super().__init__(params, cache_manager)
        self._ver = str(self.params.get("version", _DEFAULT_VERSION))
        if self._ver not in _PWT_URLS:
            supported = ", ".join(sorted(_PWT_URLS))
            raise ConnectorError(
                f"PWT version {self._ver!r} is not supported. "
                f"Supported: {supported}",
                connector_id="pwt",
            )
        self._download_url = _PWT_URLS[self._ver]
        self._cached_meta: DatasetMetadata | None = None
        self._cached_path: Path | None = None

    # ------------------------------------------------------------------
    # AbstractConnector interface
    # ------------------------------------------------------------------

    def connect(self) -> None:
        """Verify the PWT Dataverse download URL is reachable."""
        _requests = self._import_requests()
        try:
            resp = _requests.head(
                self._download_url,
                timeout=int(self.params.get("timeout", 30)),
                allow_redirects=True,
            )
            if resp.status_code not in (200, 302, 303, 307, 308):
                resp.raise_for_status()
        except Exception as exc:
            raise ConnectorError(
                f"Cannot reach PWT download URL: {self._download_url}",
                connector_id="pwt",
                cause=exc,
            ) from exc

    def download(self, *, force: bool = False) -> Path:
        """Download the PWT Excel file and write a wide-format CSV."""
        key = self.cache_key()

        if self.cache_manager is not None and not force:
            if self.cache_manager.is_cached(key):
                path, meta = self.cache_manager.retrieve(key)
                self._cached_path = path
                self._cached_meta = meta
                return path

        _requests = self._import_requests()
        self._import_openpyxl()

        timeout = int(self.params.get("timeout", 120))
        variables = self.params.get("variables") or None
        entity_col = str(self.params.get("entity_col", "country"))
        time_col = str(self.params.get("time_col", "year"))

        # Stream download to temp .xlsx
        import tempfile
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as tmp_xl:
            tmp_xl_path = Path(tmp_xl.name)
        try:
            resp = _requests.get(self._download_url, timeout=timeout, stream=True)
            resp.raise_for_status()
            with tmp_xl_path.open("wb") as fh:
                for chunk in resp.iter_content(chunk_size=65536):
                    fh.write(chunk)
        except Exception as exc:
            tmp_xl_path.unlink(missing_ok=True)
            raise ConnectorError(
                "Failed to download PWT Excel file.",
                connector_id="pwt",
                cause=exc,
            ) from exc

        # Parse Excel and write CSV
        try:
            tmp_csv_path = self._excel_to_csv(
                tmp_xl_path,
                variables=variables,
                entity_col=entity_col,
                time_col=time_col,
            )
        finally:
            tmp_xl_path.unlink(missing_ok=True)

        meta = DatasetMetadata.now(
            connector_id="pwt",
            source=f"Penn World Tables {self._ver}",
            url=self._download_url,
            version=self._ver,
            citation=_CITATION,
            params=self.params,
        )

        if self.cache_manager is not None:
            stored = self.cache_manager.store(key, tmp_csv_path, meta)
            tmp_csv_path.unlink(missing_ok=True)
            _, self._cached_meta = self.cache_manager.retrieve(key)
            self._cached_path = stored
            return stored

        self._cached_path = tmp_csv_path
        self._cached_meta = meta
        return tmp_csv_path

    def validate(self, path: Path) -> DataValidationReport:
        """Validate the downloaded wide-format CSV."""
        entity_col = str(self.params.get("entity_col", "country"))
        time_col = str(self.params.get("time_col", "year"))
        variables = list(self.params.get("variables") or [])
        required = [entity_col, time_col] + variables
        config = DataValidationConfig(
            required_columns=required,
            entity_col=entity_col,
            time_col=time_col,
            check_duplicates=True,
            check_missing_identifiers=True,
        )
        return DataValidator(config).validate_path(path)

    def metadata(self) -> DatasetMetadata:
        if self._cached_meta is None:
            raise ConnectorError(
                "No metadata available. Call download() first.",
                connector_id="pwt",
            )
        return self._cached_meta

    def cache_key(self) -> str:
        return self._make_cache_key({"version": self._ver})

    def citation(self) -> str:
        return _CITATION

    def version(self) -> str:
        return self._ver

    # ------------------------------------------------------------------
    # Public helpers
    # ------------------------------------------------------------------

    @staticmethod
    def list_versions() -> list[str]:
        """Return PWT version strings supported by this connector."""
        return sorted(_PWT_URLS)

    @property
    def download_url(self) -> str:
        """Resolved download URL for the configured PWT version."""
        return self._download_url

    # ------------------------------------------------------------------
    # Private helpers
    # ------------------------------------------------------------------

    @staticmethod
    def _import_requests() -> Any:
        try:
            import requests  # noqa: PLC0415
            return requests
        except ImportError as exc:
            raise ConnectorError(
                "PennWorldTablesConnector requires 'requests'. "
                "Install with: pip install requests",
                connector_id="pwt",
                cause=exc,
            ) from exc

    @staticmethod
    def _import_openpyxl() -> Any:
        try:
            import openpyxl  # noqa: PLC0415
            return openpyxl
        except ImportError as exc:
            raise ConnectorError(
                "PennWorldTablesConnector requires 'openpyxl'. "
                "Install with: pip install openpyxl",
                connector_id="pwt",
                cause=exc,
            ) from exc

    def _excel_to_csv(
        self,
        xlsx_path: Path,
        *,
        variables: list[str] | None,
        entity_col: str,
        time_col: str,
    ) -> Path:
        """
        Parse the PWT Excel workbook and write a tidy wide-format CSV.

        PWT Excel structure:
        - Sheet ``data``: rows = (countrycode, country, year, <variables>...)
        - First row is the header
        """
        import tempfile

        import openpyxl  # noqa: PLC0415

        try:
            wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
        except Exception as exc:
            raise ConnectorError(
                "Failed to open PWT Excel workbook.",
                connector_id="pwt",
                cause=exc,
            ) from exc

        # Find the data sheet
        sheet_names = wb.sheetnames
        data_sheet_name = next(
            (s for s in sheet_names if s.lower() in ("data", "pwt")),
            sheet_names[0] if sheet_names else None,
        )
        if data_sheet_name is None:
            raise ConnectorError(
                "PWT Excel workbook has no sheets.",
                connector_id="pwt",
            )

        ws = wb[data_sheet_name]
        rows_iter = ws.iter_rows(values_only=True)

        # Read header
        header = [str(c) if c is not None else "" for c in next(rows_iter, [])]
        if not header:
            raise ConnectorError(
                "PWT data sheet has no header row.",
                connector_id="pwt",
            )

        # Map PWT native column names to our entity/time col names
        # PWT uses 'countrycode' for ISO-3 and 'year'
        col_map: dict[str, str] = {}
        for i, col in enumerate(header):
            lc = col.lower()
            if lc in ("countrycode", "country_code", "isocode"):
                col_map[col] = entity_col
            elif lc == "year":
                col_map[col] = time_col
            else:
                col_map[col] = col

        out_header = [col_map.get(c, c) for c in header]

        # Determine which columns to keep
        if variables is not None:
            keep_set = {entity_col, time_col} | set(variables)
            keep_indices = [i for i, c in enumerate(out_header) if c in keep_set]
            out_header = [out_header[i] for i in keep_indices]
        else:
            keep_indices = list(range(len(out_header)))

        # Write to temp CSV
        with tempfile.NamedTemporaryFile(
            mode="w", suffix=".csv", delete=False, encoding="utf-8", newline=""
        ) as tmp:
            writer = csv.writer(tmp)
            writer.writerow(out_header)
            for row in rows_iter:
                out_row = [
                    "" if row[i] is None else str(row[i])
                    for i in keep_indices
                ]
                writer.writerow(out_row)
            tmp_path = Path(tmp.name)

        wb.close()
        return tmp_path
